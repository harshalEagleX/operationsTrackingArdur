"""Reports: selectors (read-only), the export job pipeline, and exporters."""

from __future__ import annotations

import csv
import io
from datetime import timedelta

import pytest

from apps.reports.exporters import CsvExporter, ExcelExporter, build_filename, get_exporter
from apps.reports.models import JobStatus, ReportJob
from apps.reports.selectors import (
    REPORT_REGISTRY,
    AllocationSelector,
    AttendanceSelector,
    BreakSelector,
    ProductivitySelector,
    ReportFilters,
    dashboard_metrics,
    get_selector,
)
from apps.reports.services import ReportService
from apps.tracking.models import SessionState, WorkSession
from core.exceptions import ValidationError
from core.timezone import now_ist, today_ist

pytestmark = pytest.mark.django_db


# ── ReportFilters ─────────────────────────────────────────────

def test_filters_from_params_splits_comma_lists():
    filters = ReportFilters.from_params({"emp_ids": "E1,E2,E3"})
    assert filters.emp_ids == ["E1", "E2", "E3"]


def test_filters_from_params_accepts_a_real_list_too():
    filters = ReportFilters.from_params({"projects": ["A", "B"]})
    assert filters.projects == ["A", "B"]


def test_filters_scoped_to_forces_an_employees_own_id(employee):
    filters = ReportFilters(emp_ids=["SOMEONE_ELSE"]).scoped_to(employee)
    assert filters.emp_ids == [employee.emp_id]


def test_filters_scoped_to_leaves_a_supervisors_filter_alone(supervisor):
    filters = ReportFilters(emp_ids=["E1", "E2"]).scoped_to(supervisor)
    assert filters.emp_ids == ["E1", "E2"]


# ── selectors ─────────────────────────────────────────────────

def test_productivity_selector_aggregates_per_employee(employee):
    WorkSession.objects.create(
        emp_id=employee.emp_id, name=employee.name, project="P",
        total_time=3600, is_started=SessionState.COMPLETED, end_time=now_ist(),
    )
    WorkSession.objects.create(
        emp_id=employee.emp_id, name=employee.name, project="P",
        total_time=1800, is_started=SessionState.COMPLETED, end_time=now_ist(),
    )

    [row] = ProductivitySelector(ReportFilters()).rows()
    assert row["total_hours"] == 1.5


def test_productivity_selector_excludes_incomplete_sessions(employee):
    WorkSession.objects.create(emp_id=employee.emp_id, project="P", is_started=SessionState.RUNNING)
    assert ProductivitySelector(ReportFilters()).rows() == []


def test_summary_selector_returns_one_row_per_session(employee):
    from apps.reports.selectors import SummarySelector

    WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(),
    )
    rows = SummarySelector(ReportFilters()).rows()
    assert len(rows) == 1


def test_break_selector_counts_overruns(employee):
    from apps.breaks.models import BreakTime

    BreakTime.objects.create(
        user_id=employee.emp_id, user_name=employee.name, break_type="Tea break 1",
        end_time=now_ist(), total_time=1000, allotted_time=300, is_overrun=True,
    )

    [row] = BreakSelector(ReportFilters()).rows()
    assert row["overruns"] == 1
    assert row["count"] == 1


def test_allocation_selector_breaks_down_by_status(employee, supervisor):
    from apps.allocations.models import AllocationStatus, BatchAllocation

    BatchAllocation.objects.create(
        allocation_id="A-1", employee_id=employee.emp_id, employee_name=employee.name,
        status=AllocationStatus.COMPLETED, quantity=100, completed_quantity=100,
    )
    BatchAllocation.objects.create(
        allocation_id="A-2", employee_id=employee.emp_id, employee_name=employee.name,
        status=AllocationStatus.PENDING, quantity=50,
    )

    [row] = AllocationSelector(ReportFilters()).rows()
    assert row["total"] == 2
    assert row["completed"] == 1
    assert row["pending"] == 1
    assert row["total_quantity"] == 150


def test_attendance_selector_aggregates_login_history(employee):
    from apps.accounts.models import LoginHistory

    LoginHistory.objects.create(
        emp_id=employee.emp_id, name=employee.name, date=today_ist(),
        login_time=now_ist(), logout_time=now_ist() + timedelta(hours=8),
    )

    rows = AttendanceSelector(ReportFilters()).rows()
    assert rows[0]["logins"] == 1


def test_date_range_filter_excludes_sessions_outside_the_window(employee):
    WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(), start_time=now_ist() - timedelta(days=30),
    )
    recent = WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(), start_time=now_ist(),
    )

    from apps.reports.selectors import SummarySelector

    filters = ReportFilters(date_from=today_ist(), date_to=today_ist())
    rows = SummarySelector(filters).rows()

    assert [r["id"] for r in rows] == [recent.id]


def test_get_selector_rejects_an_unknown_report_key():
    with pytest.raises(ValidationError):
        get_selector("not-a-real-report", ReportFilters())


def test_report_registry_matches_every_selector_class():
    for key, selector_class in REPORT_REGISTRY.items():
        instance = selector_class(ReportFilters())
        assert instance.columns  # every report declares its columns


def test_dashboard_metrics_shape(employee):
    metrics = dashboard_metrics(employee)
    assert set(metrics) == {
        "date", "active_now", "on_break_now", "sessions_today",
        "hours_today", "open_allocations", "overdue_allocations", "unacknowledged_feedback",
    }


# ── ReportService ─────────────────────────────────────────────

def test_run_returns_columns_and_rows(employee):
    WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(),
    )
    result = ReportService(actor=employee).run("productivity", {})

    assert result["report_key"] == "productivity"
    assert result["row_count"] == 1
    assert result["truncated"] is False


def test_run_scopes_an_employee_to_their_own_rows(employee, other_employee):
    WorkSession.objects.create(
        emp_id=other_employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(),
    )
    result = ReportService(actor=employee).run("productivity", {})
    assert result["row_count"] == 0


def test_run_truncates_at_the_inline_limit(employee, monkeypatch):
    import apps.reports.services as services_module

    monkeypatch.setattr(services_module, "MAX_INLINE_ROWS", 1)
    WorkSession.objects.bulk_create(
        [
            WorkSession(
                emp_id=employee.emp_id, project=f"P{i}", is_started=SessionState.COMPLETED,
                end_time=now_ist(),
            )
            for i in range(3)
        ]
    )

    result = ReportService(actor=employee).run("productivity", {})
    assert result["truncated"] is True
    assert result["row_count"] == 1


def test_queue_export_requires_an_actor():
    from core.exceptions import PermissionDeniedError

    with pytest.raises(PermissionDeniedError):
        ReportService(actor=None).queue_export("productivity", {})


def test_queue_export_rejects_an_unknown_report(employee):
    with pytest.raises(ValidationError):
        ReportService(actor=employee).queue_export("not-a-report", {})


def test_queue_export_rejects_a_bad_format(employee):
    with pytest.raises(ValidationError):
        ReportService(actor=employee).queue_export("productivity", {}, export_format="pptx")


def test_queue_export_creates_a_queued_job_and_schedules_the_task(
    django_capture_on_commit_callbacks, employee
):
    with django_capture_on_commit_callbacks(execute=True):
        job = ReportService(actor=employee).queue_export("productivity", {})

    job.refresh_from_db()
    # CELERY_TASK_ALWAYS_EAGER means the task ran synchronously inside the
    # on_commit callback — the job should already be done.
    assert job.status in (JobStatus.DONE, JobStatus.RUNNING, JobStatus.QUEUED)


def test_get_job_refuses_someone_elses_export(employee, other_employee):
    from core.exceptions import PermissionDeniedError

    job = ReportJob.objects.create(requested_by=other_employee.emp_id, report_key="productivity")
    with pytest.raises(PermissionDeniedError):
        ReportService(actor=employee).get_job(job.id)


def test_get_job_missing_id_is_not_found(employee):
    from core.exceptions import NotFoundError

    with pytest.raises(NotFoundError):
        ReportService(actor=employee).get_job(999999)


def test_mark_running_done_failed_transitions(employee):
    job = ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")
    service = ReportService(actor=employee)

    service.mark_running(job)
    job.refresh_from_db()
    assert job.status == JobStatus.RUNNING
    assert job.started_at is not None

    service.mark_done(job, stored_file=None, row_count=5)
    job.refresh_from_db()
    assert job.status == JobStatus.DONE
    assert job.row_count == 5

    service.mark_failed(job, "boom")
    job.refresh_from_db()
    assert job.status == JobStatus.FAILED
    assert job.error_message == "boom"


# ── ReportJob model ───────────────────────────────────────────

def test_job_is_finished_only_for_terminal_states(employee):
    job = ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")
    assert job.is_finished is False

    job.status = JobStatus.DONE
    assert job.is_finished is True


def test_job_duration_seconds_requires_both_timestamps(employee):
    job = ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")
    assert job.duration_seconds is None

    job.started_at = now_ist()
    job.finished_at = job.started_at + timedelta(seconds=30)
    assert 29 < job.duration_seconds < 31


# ── exporters ─────────────────────────────────────────────────

COLUMNS = [("emp_id", "Employee ID"), ("units", "Units")]
ROWS = [{"emp_id": "E1", "units": 10}, {"emp_id": "E2", "units": None}]


def test_csv_exporter_writes_a_readable_file(tmp_path):
    destination = tmp_path / "report.csv"
    CsvExporter(COLUMNS, ROWS).write(destination)

    with open(destination, encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))

    assert rows[0] == ["Employee ID", "Units"]
    assert rows[1] == ["E1", "10"]
    assert rows[2] == ["E2", ""]  # None formats to empty, not "None"


def test_excel_exporter_writes_a_readable_workbook(tmp_path):
    from openpyxl import load_workbook

    destination = tmp_path / "report.xlsx"
    ExcelExporter(COLUMNS, ROWS, title="My Report").write(destination)

    workbook = load_workbook(destination)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("Employee ID", "Units")
    assert rows[1] == ("E1", 10)


def test_get_exporter_rejects_an_unknown_format():
    with pytest.raises(ValidationError):
        get_exporter("pptx")


def test_build_filename_includes_the_report_key_and_extension():
    name = build_filename("productivity", "xlsx")
    assert name.startswith("productivity_")
    assert name.endswith(".xlsx")


def test_exporter_formats_datetimes_and_dates_consistently():
    import datetime as dt

    from apps.reports.exporters import BaseExporter

    exporter = BaseExporter([], [])
    formatted = exporter._format(dt.date(2026, 1, 5))
    assert formatted == "2026-01-05"
    assert exporter._format(None) == ""


# ── the build_report_task Celery task (runs eagerly in tests) ──

def test_build_report_task_produces_a_downloadable_file(django_capture_on_commit_callbacks, employee):
    from apps.files.models import StoredFile
    from apps.reports.tasks import build_report_task

    WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(),
    )
    job = ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")

    build_report_task(job.id)

    job.refresh_from_db()
    assert job.status == JobStatus.DONE
    assert job.row_count == 1
    assert StoredFile.objects.filter(pk=job.file_id).exists()
    assert job.file.absolute_path.exists()


def test_build_report_task_marks_failed_and_notifies_on_error(employee, monkeypatch):
    import apps.reports.selectors as selectors_module
    from apps.reports.tasks import build_report_task

    def _boom(*args, **kwargs):
        raise RuntimeError("disk full")

    # build_report_task does `from apps.reports.selectors import get_selector`
    # inside its own body on every call, so patching the source module (not
    # apps.reports.tasks, which never binds the name at module scope) is what
    # actually takes effect.
    monkeypatch.setattr(selectors_module, "get_selector", _boom)

    job = ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")
    result = build_report_task(job.id)

    job.refresh_from_db()
    assert job.status == JobStatus.FAILED
    assert "disk full" in job.error_message
    assert result["ok"] is False


def test_build_report_task_handles_a_missing_job_id():
    from apps.reports.tasks import build_report_task

    result = build_report_task(999999)
    assert result["ok"] is False


# ── HTTP ─────────────────────────────────────────────────────

def test_catalogue_endpoint_lists_every_report(as_employee):
    response = as_employee.get("/api/v1/reports/")
    keys = {row["key"] for row in response.data["data"]}
    assert keys == set(REPORT_REGISTRY)


def test_run_endpoint(as_employee, employee):
    WorkSession.objects.create(
        emp_id=employee.emp_id, project="P", is_started=SessionState.COMPLETED,
        end_time=now_ist(),
    )
    response = as_employee.post("/api/v1/reports/run/", {"report_key": "productivity"})

    assert response.status_code == 200
    assert response.data["meta"]["row_count"] == 1


def test_run_endpoint_rejects_an_invalid_date_range(as_employee):
    response = as_employee.post(
        "/api/v1/reports/run/",
        {"report_key": "productivity", "date_from": "2026-06-01", "date_to": "2026-01-01"},
    )
    assert response.status_code == 400


def test_export_endpoint_returns_202(as_employee):
    response = as_employee.post("/api/v1/reports/export/", {"report_key": "productivity"})
    assert response.status_code == 202
    assert response.data["data"]["status"] in ("queued", "running", "done")


def test_jobs_endpoint_scopes_to_the_caller(as_employee, employee, other_employee):
    ReportJob.objects.create(requested_by=employee.emp_id, report_key="productivity")
    ReportJob.objects.create(requested_by=other_employee.emp_id, report_key="productivity")

    response = as_employee.get("/api/v1/reports/jobs/")
    assert len(response.data["data"]) == 1


def test_pending_jobs_endpoint_excludes_finished_jobs(as_employee, employee):
    ReportJob.objects.create(
        requested_by=employee.emp_id, report_key="productivity", status=JobStatus.DONE
    )
    ReportJob.objects.create(
        requested_by=employee.emp_id, report_key="productivity", status=JobStatus.QUEUED
    )

    response = as_employee.get("/api/v1/reports/jobs/pending/")
    assert len(response.data["data"]) == 1


def test_metrics_endpoint_requires_a_supervisor(as_employee):
    assert as_employee.get("/api/v1/reports/metrics/").status_code == 403


def test_metrics_endpoint_for_a_supervisor(as_supervisor):
    response = as_supervisor.get("/api/v1/reports/metrics/")
    assert response.status_code == 200
    assert "sessions_today" in response.data["data"]


def test_reports_endpoints_reject_anonymous_requests(api):
    assert api.get("/api/v1/reports/").status_code in (401, 403)


def test_audit_selector_computes_accuracy(employee, supervisor):
    from apps.feedback.models import Feedback
    from apps.reports.selectors import AuditSelector

    Feedback.objects.create(
        emp_id=employee.emp_id, emp_name=employee.name, created_by=supervisor.emp_id,
        subject="x", error_count=2, sample_size=20, severity="critical",
    )

    [row] = AuditSelector(ReportFilters()).rows()
    assert row["accuracy_percent"] == 90.0
    assert row["critical_count"] == 1
    assert row["unacknowledged"] == 1


def test_audit_selector_accuracy_is_none_without_a_sample(employee, supervisor):
    from apps.feedback.models import Feedback
    from apps.reports.selectors import AuditSelector

    Feedback.objects.create(
        emp_id=employee.emp_id, created_by=supervisor.emp_id, subject="x",
    )

    [row] = AuditSelector(ReportFilters()).rows()
    assert row["accuracy_percent"] is None


def test_audit_report_over_http(as_supervisor, employee, supervisor):
    from apps.feedback.models import Feedback as FeedbackModel

    FeedbackModel.objects.create(emp_id=employee.emp_id, created_by=supervisor.emp_id, subject="x")

    response = as_supervisor.post("/api/v1/reports/run/", {"report_key": "audit"})
    assert response.status_code == 200
    assert response.data["meta"]["row_count"] == 1
