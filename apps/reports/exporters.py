"""Report exporters.

Both write to a file path rather than returning bytes: a 50,000-row export
held entirely in memory is how a Celery worker gets OOM-killed halfway
through, and ``write_only`` streaming costs nothing extra to use.
"""

from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path

from core.timezone import now_ist


class BaseExporter:
    extension = ""
    mime_type = ""

    def __init__(self, columns: list[tuple[str, str]], rows: list[dict], title: str = "Report"):
        self.columns = columns
        self.rows = rows
        self.title = title

    def write(self, destination: Path) -> Path:
        raise NotImplementedError

    def _format(self, value):
        """Normalise values that Excel and CSV disagree about."""
        if value is None:
            return ""
        if isinstance(value, dt.datetime):
            from core.timezone import to_ist

            return to_ist(value).strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, dt.date):
            return value.strftime("%Y-%m-%d")
        return value


class ExcelExporter(BaseExporter):
    extension = "xlsx"
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def write(self, destination: Path) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(self.title[:31] or "Report")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")

        from openpyxl.cell import WriteOnlyCell

        header_cells = []
        for _, label in self.columns:
            cell = WriteOnlyCell(sheet, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            header_cells.append(cell)
        sheet.append(header_cells)

        # Width has to be set before rows are streamed in write_only mode.
        for index, (_, label) in enumerate(self.columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(len(label) + 4, 14)

        for row in self.rows:
            sheet.append([self._format(row.get(key)) for key, _ in self.columns])

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return destination


class CsvExporter(BaseExporter):
    extension = "csv"
    mime_type = "text/csv"

    def write(self, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)

        # utf-8-sig: without the BOM, Excel on Windows mangles non-ASCII names.
        with open(destination, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow([label for _, label in self.columns])
            for row in self.rows:
                writer.writerow([self._format(row.get(key)) for key, _ in self.columns])

        return destination


EXPORTERS = {
    "xlsx": ExcelExporter,
    "csv": CsvExporter,
}


def get_exporter(export_format: str) -> type[BaseExporter]:
    try:
        return EXPORTERS[export_format]
    except KeyError:
        from core.exceptions import ValidationError

        raise ValidationError(
            f"Unsupported format '{export_format}'. Use one of: {', '.join(EXPORTERS)}"
        ) from None


def build_filename(report_key: str, export_format: str) -> str:
    return f"{report_key}_{now_ist():%Y%m%d_%H%M%S}.{export_format}"
