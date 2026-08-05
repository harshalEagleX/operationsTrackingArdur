"""Allocation background tasks."""

from __future__ import annotations

import logging

from celery import shared_task

from apps.allocations.models import BatchAllocation

logger = logging.getLogger("opstracking.tasks")

SLA_WARNING_HOURS = 2


@shared_task
def check_sla_breaches():
    """Warn on allocations approaching their due time. Beat, every 15 minutes.

    ``sla_notified`` keeps it to one warning per allocation rather than one
    every quarter of an hour until the deadline passes.
    """
    from apps.accounts.models import Employee
    from apps.notifications.services import NotificationService

    at_risk = BatchAllocation.objects.due_within(SLA_WARNING_HOURS).filter(sla_notified=False)

    notified = 0
    for allocation in at_risk.iterator():
        try:
            supervisor_ids = list(
                Employee.objects.supervisors().values_list("employee_id", flat=True)
            )
            NotificationService().notify(
                recipients=[allocation.employee_id, *supervisor_ids],
                notif_type="allocation.sla_breach",
                context={
                    "task_id": allocation.allocation_id,
                    "project": allocation.project,
                    "due_at": allocation.due_at.isoformat() if allocation.due_at else "",
                    "body": f"{allocation.allocation_id} is due at "
                            f"{allocation.due_at:%d %b %H:%M} and is "
                            f"{allocation.progress_percent}% complete.",
                },
                link=f"/dashboard?tab=allocations&allocation={allocation.allocation_id}",
            )
            BatchAllocation.objects.filter(pk=allocation.pk).update(sla_notified=True)
            notified += 1
        except Exception:
            logger.exception("SLA warning failed for allocation %s", allocation.allocation_id)

    if notified:
        logger.info("raised %d SLA warnings", notified)
    return {"notified": notified}


@shared_task(bind=True)
def bulk_import_allocations(self, stored_file_id: int, actor_emp_id: str):
    """Import allocations from an uploaded spreadsheet.

    Runs on the reports queue: a 5,000-row sheet takes long enough that doing
    it inside a request would hold a web worker hostage.

    Rows are validated individually and failures are collected rather than
    aborting the run — an operations team would rather import 4,980 rows and
    see a list of 20 problems than import nothing.
    """
    from openpyxl import load_workbook

    from apps.accounts.models import User
    from apps.allocations.services import AllocationService
    from apps.files.models import StoredFile

    stored = StoredFile.objects.filter(pk=stored_file_id).first()
    if not stored:
        return {"ok": False, "error": "uploaded file not found"}

    actor = User.objects.filter(emp_id=actor_emp_id).first()
    service = AllocationService(actor=actor)

    workbook = load_workbook(stored.absolute_path, read_only=True, data_only=True)
    sheet = workbook.active

    header = [str(c.value or "").strip().lower().replace(" ", "_")
              for c in next(sheet.iter_rows(min_row=1, max_row=1))]

    imported, errors = 0, []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {key: value for key, value in zip(header, row, strict=False)
                  if key and value is not None}
        if not record.get("allocation_id"):
            continue
        try:
            service.create(
                {
                    "allocation_id": str(record["allocation_id"])[:50],
                    "employee_id": str(record.get("employee_id", ""))[:20],
                    "employee_name": str(record.get("employee_name", ""))[:100],
                    "project": str(record.get("project", ""))[:150],
                    "client_code": str(record.get("client_code", ""))[:50],
                    "work_type": str(record.get("work_type", ""))[:100],
                    "batch": str(record.get("batch", ""))[:100],
                    "order_id": str(record.get("order_id", ""))[:100],
                    "quantity": int(record.get("quantity", 0) or 0),
                }
            )
            imported += 1
        except Exception as exc:
            errors.append({"row": row_number, "error": str(exc)[:200]})

    workbook.close()

    logger.info("bulk import: %d rows imported, %d failed", imported, len(errors))
    return {"ok": True, "imported": imported, "failed": len(errors), "errors": errors[:50]}
